import openpyxl


def date_to_excel_file(date) :
    # เรียก Excel file + 'sheet1'
    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook['Sheet2']
    # หาบรรทัดล่าสุด
    last_row = sheet.max_row
    # เปลี่ยนชื่อ column เป็น index  
    col_num = openpyxl.utils.column_index_from_string('A')
    # append date to new row
    sheet.cell(row = last_row + 1 , column = col_num).value = date
    workbook.save('data.xlsx')

def excel_file(all_data_to_excel):
    #เรียก excel
    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook['Sheet2']

    #หาบรรทัดสุดท้าย
    last_row = sheet.max_row

    #ชื่อคอลั่มในการเรียกไปลงใน excel
    columns_dict = {                               'Balance (before trade)': 'B', 
                    'Balance (after trade)': 'C',  'Check signal' : 'D',
                    'Buy signal' : 'E' ,           'Sell signal' : 'F',
                    'Sell value' : 'G' ,           'Limit order' : 'H',
                    'Stop loss percent' : 'I' ,    'Trade volume' : 'J',
                    'EMA fast' : 'K' ,             'EMA slow' : 'L' ,
                    'Time frame' : 'M' ,           'Cancel litmit order' : 'N',
                    'Dust asset' : 'O' ,           'Dust to BNB' : 'P',
                    }

# Date	/ Balance (before trade) / Balance (after trade) / Check signal / Buy signal / Sell signal / Sell value	 / Limit order / Stop loss percent	
# Trade volume	/ EMA fast	/ EMA slow / Time frame / cancel litmit order
    
    
    # loop dict กับ list ใน list ที่เราเก็บมา
    for column_names , each_values in zip(columns_dict.keys() , all_data_to_excel.values()) :
    
    #เปลี่ยน input เป็นตัวอักษรจาก dict เพื่อให้ลงได้
        column_letter = columns_dict[column_names]
        column_number = openpyxl.utils.column_index_from_string(column_letter)
        # [] list ว่าง แทนค่าด้วย nan
        if len(each_values) == 0:
            
            sheet.cell(row = last_row , column = column_number).value = 'nan'
        
        # แทนค่า ... ด้วย nan
        elif each_values == "[...]" :
            sheet.cell(row = last_row , column = column_number).value = 'nan'
        
        # มีอันลบ [] ออกแล้วใส่ค่าปกติ
        elif len(each_values) == 1:
            to_excel = f'{each_values}'
            to_excel = to_excel.replace("[","")
            to_excel = to_excel.replace("]","")
            sheet.cell(row = last_row , column = column_number).value = to_excel
        
        # มี 2ตัวขึ้นไป ใช้ join method
        else:
            sheet.cell(row = last_row , column = column_number).value = ', '.join(each_values)
        
    # save
    workbook.save('data.xlsx')