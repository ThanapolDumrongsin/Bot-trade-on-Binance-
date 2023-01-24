import ccxt
import pandas as pd
import pandas_ta as ta
import time
from binance.client import Client 
import mplfinance as mpf
import numpy as np
import openpyxl
from openpyxl.utils import column_index_from_string

class setup_condition :
    # setup Client and exchange
    def __init__(self, apikey , secret ):
        self.apikey = apikey
        self.secret = secret
        self.my_account = Client(apikey, secret)
        self.exchange = ccxt.binance({'apiKey': apikey, 'secret': secret, 'enableRateLimit': True})
    
    # check your balance
    def balance(self,money_trade):
        balance = self.my_account.get_asset_balance(asset=money_trade)['free']
        balance = int(float(balance))
        return balance

    # price data from exchange
    def pricedata(self,pair,timeframe,limit):
        #ดึงข้อมูลจาก Exchange
        df_ohlcv = self.exchange.fetch_ohlcv(pair, timeframe = timeframe , limit = limit)                  
        #ดึงมาเรียงในแบบ columns (ตาราง)
        data_price = pd.DataFrame(df_ohlcv, columns=['Datetime', 'Open', 'High', 'Low', 'Close', 'Volume'])   
        
        #เปลี่ยนรูปแบบวันที่ให้คนอ่านได้
        data_price['Datetime'] = pd.to_datetime(data_price['Datetime'], unit='ms')  
        data_price.set_index('Datetime', inplace=True)                                                     
        
        #เพิ่ม pair ต่อท้ายข้อมูล
        data_price['Pair'] = pair
        return data_price

    # calculate ema and add to data chart
    def EMA(data_price , ema_fast , ema_slow):
        ema_fast_data = round(data_price.ta.ema(ema_fast),6)
        ema_slow_data = round(data_price.ta.ema(ema_slow),6)
        ema = pd.concat([data_price, ema_fast_data, ema_slow_data], axis=1)                 #รับ EMA มาต่อท้ายข้อมูล
                                                                   
        
        ema_back = len(ema)
        # Ema เร็ว ย้อนหลัง 3 วัน  (ที่ไม่ใช้ของวันนี้เพราะเทรดบน tf day)
        ema_fast_1 = ema['EMA_12'][ema_back - 2]
        ema_fast_2 = ema['EMA_12'][ema_back - 3]
        ema_fast_3 = ema['EMA_12'][ema_back - 4]
        # Ema ช้า ย้อนหลัง 3 วัน (ที่ไม่ใช้ของวันนี้เพราะเทรดบน tf day)
        ema_slow_1 = ema['EMA_26'][ema_back - 2]
        ema_slow_2 = ema['EMA_26'][ema_back - 3]
        ema_slow_3 = ema['EMA_26'][ema_back - 4]
        return ema , ema_fast_1 ,ema_fast_2 , ema_fast_3 ,ema_slow_1 ,ema_slow_2 , ema_slow_3

    
    def marketprice_and_sl(self,pair , stop_loss_percent):
        marketprice = self.my_account.get_symbol_ticker(symbol = pair) 
        marketprice = float(marketprice[ 'price' ])
        # calculate stoploss price
        stop_loss_percent = (100 - stop_loss_percent) /100
        stop_loss_percent = round(stop_loss_percent,2)
        sl_price =  float(marketprice)
        # stoploss relate to your desire stoploss percent
        sl_price =  sl_price * stop_loss_percent 
        return marketprice , sl_price
    
    # round price and amount for binance (each coin have different decimal for price and amount)
    def round_sell_and_price(sl_price):
        amount_sell_decimal = None
        sell_price_decimal = None
        # round เพื่อส่งคำสั่งซื้อขายแล้วไม่เกิด Error + ทำให้ไม่ขาดทุนในบางครั้ง
        if sl_price > 1000 :
            sell_price_decimal = None
            amount_sell_decimal = 5
        elif sl_price > 100 :
            sell_price_decimal = 1
            amount_sell_decimal = 3
        elif sl_price > 10 :
            sell_price_decimal = 2
            amount_sell_decimal = 1
        elif sl_price > 0 :
            sell_price_decimal = 3
            amount_sell_decimal = 1
        elif sl_price < 0 :
            sell_price_decimal = 3
            amount_sell_decimal = None
        
        return amount_sell_decimal , sell_price_decimal
    
    # check your coin balance
    def sell_amount(self,each_coin):
        coin_balance = self.my_account.get_asset_balance(asset= each_coin)['free']
        sell_amount = float(coin_balance)
        return sell_amount
    
    #check signal
    def check_signal(pair , ema_fast_1 , ema_slow_1 , ema_fast_2 , ema_slow_2 ,check_signal_excel):
        
        signal = "No signal"
        
        if ema_fast_1 > ema_slow_1 and ema_fast_2 < ema_slow_2 :   
            check_signal_result = f"Check {pair} for buy signal next bar.\n"
            print(check_signal_result)
            
            check_signal_excel.append(check_signal_result)
            
            signal = "check"

        else :
            print(f"Check signal {pair} = {signal}")
        return signal
        
    # buy signal
    def buy_signal(self,pair,ema_fast_1,ema_fast_2,ema_fast_3,ema_slow_2,ema_slow_3,BUSDbalance,trade_volume,marketprice
    ,each_coin,amount_sell_decimal,sl_price,sell_price_decimal , buy_signal_excel):
        signal = "No signal"
        
        if ema_fast_1 > ema_fast_2 > ema_slow_2 and ema_fast_3 < ema_slow_3 :
            BUSDbalance = self.my_account.get_asset_balance(asset='BUSD')['free']
            BUSDbalance = int(float(BUSDbalance))
            if BUSDbalance > trade_volume :
                self.my_account.create_order(                                                                
                    symbol = pair,
                    side = Client.SIDE_BUY,
                    type = Client.ORDER_TYPE_MARKET,
                    quoteOrderQty = trade_volume
                    )
                buy_result = f"Buy signal of {pair} at price {marketprice}$."
                print(buy_result)
                
                # รอ 2 วิ แล้วตั้งคำสั่ง stop loss order
                time.sleep(2)
                
                # Check ว่าเรามีเหรียญอยู่เท่าไหร่
                balance = self.my_account.get_asset_balance(asset= each_coin)['free']
                balance_round = round(balance , amount_sell_decimal)
                sl_price_round = round(sl_price , sell_price_decimal)
                sell_amount = float(balance)
                #ลดปริมาณการขายจากที่มีเล็กน้อยเพื่อป้องกัน unexpected error
                sell_amount = sell_amount / 1.001

                # SL order (limit)
                self.my_account.create_order(                                                                           
                    symbol = pair,
                    side = Client.SIDE_SELL,
                    type = Client.ORDER_TYPE_STOP_LOSS_LIMIT,
                    timeInForce = "GTC",
                    price = round(sl_price , sell_price_decimal),
                    quantity = round(sell_amount , amount_sell_decimal),
                    stopPrice = round(sl_price , sell_price_decimal)
                    )
                sl_result = (f" , amount of {each_coin} = {balance_round} coin, Stop loss price at {sl_price_round}.\n")
                print(sl_result)
                signal = "buy_signal"

                buy_signal_excel.append(buy_result+sl_result)
                
            else :
                buy_but_no_money_result = f"Buy signal for {each_coin}, But your account balance have {BUSDbalance}$. \n not enough for your trading volume\n"
                print(buy_but_no_money_result)
                signal = "buy_signal_but_no_money"
                buy_signal_excel.append(buy_but_no_money_result)
        else :
            print(f"Buy_signal {pair} = {signal}")
        
        return signal
        

    def cancel_limit(self ,pair, ema_fast_1,ema_slow_1,ema_fast_2,ema_slow_2 ,cancel_limit_excel ):
        # Ema เร็ว ตัดลง Ema ช้า
        if ema_fast_1 < ema_slow_1 and ema_fast_2 > ema_slow_2 :
        # รับ Id order
            get_limit_order = self.my_account.get_open_orders(symbol= pair)
            for order in get_limit_order:
                order_id = order['orderId']
                
                # ยกเลิก Limit order ถ้ามี limit order อยู่
                if order['orderId'] == order_id and order['symbol'] == pair and order['side'] == 'SELL' and order['status'] == 'NEW':
                    self.my_account.cancel_order(symbol = pair , orderId = order_id)
                    
                    cancel_litmit_result = f"Cancel limit order {pair}.\n"
                    
                    cancel_limit_excel.append(cancel_litmit_result)
                    print(cancel_litmit_result)
            
                #ถ้าไม่มี limit order = ผ่าน
                else :
                    pass
    
    def sell_signal (self, pair , each_coin , ema_fast_1 , ema_slow_1 , ema_fast_2 , ema_slow_2 , marketprice ,
    sell_price_decimal , amount_sell_decimal , sell_signal_excel , sell_value_excel):
        signal = "No signal"
        #รอ 2 วิ แล้วส่งคำสั่งขาย
        time.sleep(2)
        # Check ว่าเรามีเหรียญอยู่เท่าไหร่
        coin_balance = self.my_account.get_asset_balance(asset= each_coin)['free']
        sell_amount = float(coin_balance)
        #ลดปริมาณการขายจากที่มีเล็กน้อยเพื่อป้องกัน unexpected error
        sell_amount = sell_amount / 1.001
        
        # จาก def marketprice_and_sl
        sell_value = marketprice

        # หาว่ามูลค่าเหรียญใน บช มีค่าเท่าไหร่
        sell_value = round(sell_value * sell_amount , sell_price_decimal)
        if ema_fast_1 < ema_slow_1 and ema_fast_2 > ema_slow_2 :
        # ถ้าราคาเกิน 10$ จะส่งคำสั่งขาย
            if sell_value > 10 :
                # ส่งคำสั่งขาย
                self.my_account.create_order(                                                                
                    symbol = pair,
                    side = Client.SIDE_SELL,
                    type = Client.ORDER_TYPE_MARKET,
                    quantity= round(sell_amount,amount_sell_decimal)
                    )
                sell_result = f"sell {pair} for {sell_value} $.\n"
                print(sell_result)
                signal = "sell_signal"
                sell_signal_excel.append(sell_result)
                sell_value_excel.append(f"{pair} = {sell_value}$.")
            
            # Value ไม่ถึง 10 $
            else : 
                sell_but_no_coin_result = f"{pair} is down trend but your {pair} value < 10$ can't sell.\n"
                print(sell_but_no_coin_result)
                signal = "sell_signal_but_no_coin"
                sell_signal_excel.append(sell_but_no_coin_result)
        else :
            print(f"Sell_signal {pair} = {signal} \n")
        
        return signal , sell_value

    def sell_bnb(self,sell_value_excel,money_trade):
        coin_balance = self.my_account.get_asset_balance(asset= 'BNB')['free']
        sell_amount = float(coin_balance)
        #ลดปริมาณการขายจากที่มีเล็กน้อยเพื่อป้องกัน unexpected error
        sell_amount = sell_amount / 1.001
        marketprice = self.my_account.get_symbol_ticker(symbol =f'BNB{money_trade}') 
        marketprice = float(marketprice[ 'price' ])
        # จาก def marketprice_and_sl
        sell_value = marketprice
        # หาว่ามูลค่าเหรียญใน บช มีค่าเท่าไหร่
        sell_value = round(sell_value * sell_amount , 1)
        if sell_value > 10 :
                # ส่งคำสั่งขาย
            self.my_account.create_order(                                                                
                symbol = 'BNB',
                side = Client.SIDE_SELL,
                type = Client.ORDER_TYPE_MARKET,
                quantity= round(sell_amount,3)
                )
            sell_result = f"Sell BNB for {sell_value}$.\n"
            print(sell_result)
            sell_value_excel.append(sell_result)
    
    def graph_pic(self,price_chart_with_ema , limit ,pair , signal , sl_price ,marketprice,each_coin,trade_volume,sell_value , sell_price_decimal):
        graph_pic = price_chart_with_ema
        add_ema_data_12 = price_chart_with_ema['EMA_12']
        add_ema_data_26 = price_chart_with_ema['EMA_26']
        add_fill_between_up = dict(y1=add_ema_data_12.values , y2=add_ema_data_26.values , where = add_ema_data_12 > add_ema_data_26 , color="#93c47d" , alpha=0.3 , interpolate=True)
        add_fill_between_down = dict(y1=add_ema_data_12.values , y2=add_ema_data_26.values , where = add_ema_data_12 < add_ema_data_26 , color="#e06666" , alpha=0.3 , interpolate=True)
        fill_between = [add_fill_between_up , add_fill_between_down]

        up_trend_sign = []
        for day in range(len(price_chart_with_ema)):
            if  price_chart_with_ema['EMA_26'][day] > price_chart_with_ema['EMA_12'][day] and price_chart_with_ema['EMA_12'][day+1] > price_chart_with_ema['EMA_26'][day+1] :
                up_trend_sign.append(add_ema_data_26[day+1]*0.95)
            else:
                up_trend_sign.append(np.nan)

        down_trend_sign = [np.nan]
        for i in range (len(price_chart_with_ema)-1) :
            if price_chart_with_ema['EMA_26'][i+1] > price_chart_with_ema['EMA_12'][i+1] and price_chart_with_ema['EMA_12'][i] > price_chart_with_ema['EMA_26'][i] :
                down_trend_sign.append(add_ema_data_26[i+1]*1.05)
            else :
                down_trend_sign.append(np.nan)

        add_all_data = [mpf.make_addplot(add_ema_data_12 , color = 'lime'),
                        mpf.make_addplot(add_ema_data_26 , color = 'r'),
                        mpf.make_addplot(up_trend_sign ,type='scatter',markersize=200,marker='^',color="lime"),
                        mpf.make_addplot(down_trend_sign ,type='scatter',markersize=200,marker='v',color="r")
                    ]

        # ตำแหน่ง และจำนวน hline+สี
        if signal == "check" :
            hline = dict( hlines = [ marketprice] , colors=['g'] , linestyle = '-.' ,)
            print(f"Create graph for {pair}")
        elif signal == "buy_signal":
            hline = dict( hlines = [sl_price , marketprice] , colors=['r', 'g'] , linestyle = '-.' ,)
            print(f"Create graph for {pair}")
        elif signal == "buy_signal_but_no_money":
            hline = dict( hlines = [sl_price , marketprice] , colors=['#e8e046', '#e8e046'] , linestyle = '-.' ,)
            print(f"Create graph for {pair}")
        elif signal == "sell_signal" :
            hline = dict( hlines = [ marketprice] , colors=['r'] , linestyle = '-.' ,)
            print(f"Create graph for {pair}")
        elif signal == "sell_signal_but_no_coin" :
            hline = dict( hlines = [ marketprice] , colors=['r'] , linestyle = '-.' ,)
            print(f"Create graph for {pair}")      
        
        # create graph
        fig,axlist = mpf.plot(graph_pic ,hlines = hline, title = f"{price_chart_with_ema.index[-1]} \n {pair}." ,volume = True , addplot = add_all_data , type='candle' , figscale=2 , figsize=(20,10),
            style = 'binance' , tight_layout=True , fill_between = fill_between , returnfig = True)
        
        # text position
        x_position_for_text = limit *0.75
        stop_loss_percent = (100 - stop_loss_percent) /100
        stop_loss_percent = round(stop_loss_percent,2)
        
        # calculate how much money gonna lose if stoploss trigger 
        loss_calculation = trade_volume - (trade_volume * stop_loss_percent)
        
        # how much profit you get  (sell_value from def sell_signal)
        profit =  sell_value - trade_volume
       
        # Text for Horizontal line , each signal
        if signal == "check" :
            axlist[0].text(x_position_for_text, marketprice*1.03 , f"Check buy signal Tomorrow (7.00 AM).", fontweight='bold' ,color = 'g')
        
        elif signal == "buy_signal":
            axlist[0].text(x_position_for_text, sl_price*1.03 , f"Stop loss at {round(sl_price , sell_price_decimal)}$ \n Loss if Stop loss = {loss_calculation}$.", fontweight='bold' ,color = 'r')
            axlist[0].text(x_position_for_text, marketprice*1.03 , f"Buy '{pair}' at {marketprice}$ for {trade_volume}.", fontweight='bold' ,color = 'g')
        
        elif signal == "buy_signal_but_no_money" :
            axlist[0].text(x_position_for_text, sl_price*1.03 , f"if buy Stop loss at {round(sl_price , sell_price_decimal)}$ \n Loss if Stop loss = {loss_calculation}$.", fontweight='bold' ,color = '#e8e046')
            axlist[0].text(x_position_for_text, marketprice*1.03 , f"Buy signal for '{pair}' at {marketprice}$ \n But your balance is less than trade volume.", fontweight='bold' ,color = '#e8e046')
        
        elif signal == "sell_signal" :
            axlist[0].text(x_position_for_text, marketprice*1.03 , f"sell signal for '{pair}' at {marketprice}$ \n Your profit is {profit}$.", fontweight='bold' ,color = '#e8e046')
        
        elif signal == "sell_signal_but_no_coin" :
            axlist[0].text(x_position_for_text, marketprice*1.03 , f"sell signal for '{pair}' at {marketprice}$ \n But your {each_coin} value is less than 10$.", fontweight='bold' ,color = '#e8e046')
        
        # folder เก็บภาพ
        path = "graph_pic"

        # แปลงชื่อ file ให้อ่านง่าย + save
        filename = f'{pair}{str(price_chart_with_ema.index[-1]).replace(":","")}.png'
        filename = filename.replace("000000","") 
        fig.savefig(f"{path}/{filename}")


    def check_order(self , pair , sell_price_decimal , check_limit_order_excel):
        check_order= self.my_account.get_open_orders(symbol= pair)
        for order in check_order:
            order_id = order['orderId']
            
            # Check limit order (daily)
            if order['orderId'] == order_id and order['symbol'] == pair and order['side'] == 'SELL' and order['status'] == 'NEW':
                
                price = float(order['price']) 
                amount = float(order['origQty'])

                value_calculation = round(price * amount , sell_price_decimal)
                check_order_result = f"limit order {pair} at price {price}$ value = {value_calculation}$.\n"
                print(check_order_result)
                check_limit_order_excel.append(check_order_result)
                return check_order_result
    
    # check for coin that able to dust to BNB
    def asset_that_can_dust(self , asset_result_excel) :
        asset_that_can_dust = []
        get_dust_assets = self.my_account.get_dust_assets()
        for asset in get_dust_assets['details']:
            result = f"{asset['asset']}  can convert into {asset['toBNB']} BNB.\n"
            print(result)
            asset_that_can_dust.append(asset['asset'])
            asset_result_excel.append(result)
        
        total_BNB = f"\nTotal convert amount BNB = {get_dust_assets['totalTransferBNB']}."
        print(total_BNB)
        asset_result_excel.append(asset_result_excel)
        return asset_that_can_dust

    # dust to BNB  
    def dust_to_BNB(self , asset_that_can_dust , dust_to_BNB_excel):
        
        # if we have more than 1 coin to dust
        if len(asset_that_can_dust) >= 1 :
            transfer_dust = self.my_account.transfer_dust(asset = ", ".join(asset_that_can_dust))
        
            transfer_dust_result = transfer_dust['totalTransfered']
            print(transfer_dust_result)
            dust_to_BNB_excel.append(transfer_dust_result)
        
        # only one coin to dust 
        elif len(asset_that_can_dust ) == 1 :
            transfer_dust = self.my_account.transfer_dust(asset = asset_that_can_dust)
            transfer_dust_result = transfer_dust['totalTransfered']
            print(transfer_dust_result)
            dust_to_BNB_excel.append(transfer_dust_result)
            
        time.sleep(1)
        
    


def date_to_excel_file(value) :
    # เรียก Excel file + 'sheet1'
    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook['Sheet2']
    # หาบรรทัดล่าสุด
    last_row = sheet.max_row
    # เปลี่ยนชื่อ column เป็น index  
    col_num = openpyxl.utils.column_index_from_string('A')
    # append date to new row
    sheet.cell(row = last_row + 1 , column = col_num).value = value
    workbook.save('data.xlsx')




def excel_file(all_data_to_excel):
    #เรียก excel
    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook['Sheet2']

    #หาบรรทัดสุดท้าย
    last_row = sheet.max_row

    #ชื่อคอลั่มในการเรียกไปลงใน excel
    columns_dict = {                               'Balance (before trade)': 'B', 
                    'Balance (after trade)': 'C',  'Check signal' : 'D',
                    'Buy signal' : 'E' ,           'Sell signal' : 'F',
                    'Sell value' : 'G' ,           'Limit order' : 'H',
                    'Stop loss percent' : 'I' ,    'Trade volume' : 'J',
                    'EMA fast' : 'K' ,             'EMA slow' : 'L' ,
                    'Time frame' : 'M' ,           'Cancel litmit order' : 'N',
                    'Dust asset' : 'O' ,           'Dust to BNB' : 'P',
                    }

# Date	/ Balance (before trade) / Balance (after trade) / Check signal / Buy signal / Sell signal / Sell value	 / Limit order / Stop loss percent	
# Trade volume	/ EMA fast	/ EMA slow / Time frame / cancel litmit order
    
    # ค่า list ใน list จาก parameter
    list_values = all_data_to_excel
    
    # loop dict กับ list ใน list ที่เราเก็บมา
    for column_names , each_values in zip(columns_dict.keys() , list_values) :
    
    #เปลี่ยน input เป็นตัวอักษรจาก dict เพื่อให้ลงได้
        column_letter = columns_dict[column_names]
        column_number = openpyxl.utils.column_index_from_string(column_letter)
        # [] list ว่าง แทนค่าด้วย nan
        if len(each_values) == 0:
            
            sheet.cell(row = last_row , column = column_number).value = 'nan'
        
        # แทนค่า ... ด้วย nan
        elif each_values == "..." :
            sheet.cell(row = last_row , column = column_number).value = 'nan'
        
        # มีอันลบ [] ออกแล้วใส่ค่าปกติ
        elif len(each_values) == 1:
            to_excel = f'{each_values}'
            to_excel = to_excel.replace("[","")
            to_excel = to_excel.replace("]","")
            sheet.cell(row = last_row , column = column_number).value = to_excel
        
        # มี 2ตัวขึ้นไป ใช้ join method
        else:
            sheet.cell(row = last_row , column = column_number).value = ', '.join(each_values)
        
    # save
    workbook.save('data.xlsx')
   
   
   
   



